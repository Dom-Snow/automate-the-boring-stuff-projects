import openpyxl

from openpyxl.styles import Font
font_title = Font(bold=True)

math_tab = openpyxl.Workbook()
sheet1 = math_tab.active

N = 7

for i in range(1, N):
    sheet1.cell(row=1, column=i+1).value = i
    sheet1.cell(row=1, column=i + 1).font = font_title
    sheet1.cell(row=i+1, column=1).value = i
    sheet1.cell(row=i + 1, column=1).font = font_title

for i in range(2, N+1):
    for j in range(2, N+1):
        sheet1.cell(row=i, column=j).value = (i-1) * (j-1)
        
math_tab.save("math_tab.xlsx")
