import openpyxl

items = openpyxl.load_workbook("sample.xlsx")
sheet = items.active

invert = openpyxl.Workbook()
sheet1 = invert.active

R = sheet.max_row

for i in range(1, R):
    sheet1.cell(row=1, column=i).value = sheet.cell(row=i, column=1).value
    sheet1.cell(row=2, column=i).value = sheet.cell(row=i, column=2).value

invert.save("invert.xlsx")
