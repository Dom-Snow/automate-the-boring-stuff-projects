import openpyxl

N = 5 # Row after which we want to insert R blank rows
R = 3 # Number of blank rows we want to insert

wb = openpyxl.load_workbook('your_file.xlsx')
sheet = wb.active

for r in range(sheet.max_row, N, -1):
    for c in range(1, sheet.max_column + 1):
        sheet.cell(row=r+R, column=c).value = sheet.cell(row=r, column=c).value

for r in range(N+1, N+R+1):
    for c in range(1, sheet.max_column + 1):
        sheet.cell(row=r, column=c).value = None

wb.save('iterbyrow.xlsx')
